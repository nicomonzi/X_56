#!/usr/bin/env python3
"""Create publication-quality plots of the X-56A spanwise airfoil sections."""

from pathlib import Path

import matplotlib as mpl
import matplotlib.pyplot as plt
import numpy as np
from openpyxl import load_workbook


ROOT = Path(__file__).resolve().parent
SOURCE = ROOT / "X-56A_Airfoil_Sections.xlsx"
OUTPUT = ROOT / "plots"

# Full local chord extracted from the left-half CAERO1 planform in
# ../TRIM/X56A-NASTRAN-GEOM.dat.  At control-surface stations the forward
# CAERO1 chord and the 5.563-in body-flap chord are combined.
PLANFORM_Y_IN = np.asarray([0.000, 9.350, 19.021, 21.203, 25.665, 45.732, 50.000])
PLANFORM_CHORD_IN = np.asarray(
    [92.453, 72.682, 52.232, 49.490, 43.884, 27.484, 23.996]
)
PLANFORM_X_LE_IN = np.asarray(
    [100.000, 115.914, 132.417, 134.263, 138.038, 146.376, 148.149]
)
PLANFORM_Z_LE_IN = np.asarray(
    [100.861, 101.481, 102.299, 102.410, 102.638, 103.893, 104.160]
)


def read_sections(path: Path):
    """Read the chordwise coordinates and all airfoil stations from the workbook."""
    sheet = load_workbook(path, data_only=True, read_only=True).active
    x_over_c = np.asarray(
        [value for value in next(sheet.iter_rows(min_row=8, max_row=8, values_only=True))[2:] if value is not None],
        dtype=float,
    )

    sections = []
    row = 10
    while row <= sheet.max_row:
        name = sheet.cell(row, 1).value
        station_label = sheet.cell(row + 1, 1).value if row + 1 <= sheet.max_row else None
        if name and station_label == "Spanwise, y=":
            y_in = float(sheet.cell(row + 1, 2).value)
            upper = np.asarray(
                [sheet.cell(row + 2, column).value for column in range(3, 3 + len(x_over_c))],
                dtype=float,
            )
            lower = np.asarray(
                [sheet.cell(row + 3, column).value for column in range(3, 3 + len(x_over_c))],
                dtype=float,
            )
            sections.append(
                {
                    "name": str(name).strip(),
                    "y_in": y_in,
                    "r_in": abs(y_in),
                    "chord_in": float(
                        np.interp(abs(y_in), PLANFORM_Y_IN, PLANFORM_CHORD_IN)
                    ),
                    "x_le_in": float(
                        np.interp(abs(y_in), PLANFORM_Y_IN, PLANFORM_X_LE_IN)
                    ),
                    "z_le_in": float(
                        np.interp(abs(y_in), PLANFORM_Y_IN, PLANFORM_Z_LE_IN)
                    ),
                    "upper": upper,
                    "lower": lower,
                }
            )
            row += 7
        else:
            row += 1

    # The workbook describes the left semispan (negative y).  Sort by distance
    # from the aircraft centreline to make the colour progression unambiguous.
    sections = [
        section for section in sections
        if not section["name"].casefold().startswith("bf")
    ]
    sections.sort(key=lambda item: item["r_in"])
    return x_over_c, sections


def configure_style():
    mpl.rcParams.update(
        {
            "font.family": "DejaVu Sans",
            "font.size": 10,
            "axes.titlesize": 15,
            "axes.labelsize": 11,
            "axes.titleweight": "bold",
            "axes.spines.top": False,
            "axes.spines.right": False,
            "axes.grid": True,
            "grid.color": "#c9d1d9",
            "grid.linewidth": 0.65,
            "grid.alpha": 0.55,
            "legend.frameon": False,
            "figure.facecolor": "white",
            "axes.facecolor": "#fbfcfe",
            "savefig.facecolor": "white",
        }
    )


def station_colours(sections):
    # Plasma remains strongly saturated throughout this clipped range and avoids
    # the pale-yellow end of several sequential maps, which is weak on white.
    values = np.linspace(0.03, 0.82, len(sections))
    return mpl.colormaps["plasma"](values)


def save_figure(fig, stem):
    OUTPUT.mkdir(exist_ok=True)
    fig.savefig(OUTPUT / f"{stem}.png", dpi=300, bbox_inches="tight")
    fig.savefig(OUTPUT / f"{stem}.pdf", bbox_inches="tight")
    plt.close(fig)


def plot_superimposed(x, sections, colours):
    fig, ax = plt.subplots(figsize=(12.0, 5.8), constrained_layout=True)

    for section, colour in zip(sections, colours):
        chord = section["chord_in"]
        x_in = section["x_le_in"] + x * chord / 100.0
        upper_in = section["z_le_in"] + section["upper"] * chord / 100.0
        lower_in = section["z_le_in"] + section["lower"] * chord / 100.0
        label = section["name"]
        ax.plot(x_in, upper_in, color=colour, lw=2.5, label=label)
        ax.plot(x_in, lower_in, color=colour, lw=2.5)

    ax.set(
        xlabel="$x$ (in)",
        ylabel="$z$ (in)",
        xlim=(96, 196),
    )
    ax.set_aspect("equal", adjustable="box")
    save_figure(fig, "airfoil_sections_superimposed")


def plot_planform_positions(sections, colours):
    """Plot each section at its true chordwise and spanwise planform position."""
    fig, ax = plt.subplots(figsize=(12.0, 7.2), constrained_layout=True)

    for section, colour in zip(sections, colours):
        if section["name"].casefold() == "body (center)":
            y = 0.0
            x_le = float(PLANFORM_X_LE_IN[0])
            x_te = x_le + float(PLANFORM_CHORD_IN[0])
        else:
            y = section["r_in"]
            x_le = section["x_le_in"]
            x_te = x_le + section["chord_in"]
        ax.plot(
            [x_le, x_te],
            [y, y],
            color=colour,
            lw=3.0,
            solid_capstyle="round",
            label=section["name"],
        )
        ax.scatter(x_le, y, color=colour, s=32, zorder=3, edgecolor="white", linewidth=0.6)

    x_te_knots = PLANFORM_X_LE_IN + PLANFORM_CHORD_IN
    ax.plot(
        PLANFORM_X_LE_IN,
        PLANFORM_Y_IN,
        color="#111111",
        lw=1.8,
        ls="--",
        label="Leading edge",
        zorder=2,
    )
    ax.plot(
        x_te_knots,
        PLANFORM_Y_IN,
        color="#555555",
        lw=1.5,
        ls=":",
        label="Trailing edge",
        zorder=2,
    )

    ax.set(
        xlabel="$x$ (in)",
        ylabel="$y$ (in)",
        xlim=(96, 196),
        ylim=(-2, 53),
    )
    ax.invert_yaxis()
    ax.set_aspect("equal", adjustable="box")
    ax.legend(
        loc="center left",
        bbox_to_anchor=(1.015, 0.5),
        fontsize=8.7,
    )
    save_figure(fig, "airfoil_sections_planform_positions")


def plot_sections_3d(x, sections, colours):
    """Show the dimensional sections at their true spanwise positions."""
    fig = plt.figure(figsize=(12.8, 3.8))
    ax = fig.add_subplot(111, projection="3d")

    le_x, le_y, le_z = [], [], []
    te_x, te_y, te_z = [], [], []
    for section, colour in zip(sections, colours):
        chord = section["chord_in"]
        x_in = section["x_le_in"] + x * chord / 100.0
        upper_in = section["z_le_in"] + section["upper"] * chord / 100.0
        lower_in = section["z_le_in"] + section["lower"] * chord / 100.0
        y_in = np.full_like(x_in, section["r_in"])

        ax.plot(x_in, y_in, upper_in, color=colour, lw=1.25)
        ax.plot(x_in, y_in, lower_in, color=colour, lw=1.25)
        ax.plot(
            [x_in[-1], x_in[-1]],
            [section["r_in"], section["r_in"]],
            [lower_in[-1], upper_in[-1]],
            color=colour,
            lw=1.25,
        )

        le_x.append(x_in[0])
        le_y.append(section["r_in"])
        le_z.append(0.5 * (upper_in[0] + lower_in[0]))
        te_x.append(x_in[-1])
        te_y.append(section["r_in"])
        te_z.append(0.5 * (upper_in[-1] + lower_in[-1]))

    ax.plot(le_x, le_y, le_z, color="#111111", lw=1.0, ls="--", label="Leading edge")
    ax.plot(te_x, te_y, te_z, color="#555555", lw=0.9, ls=":", label="Trailing edge")
    ax.set(
        xlabel="$x$ (in)",
        ylabel="$y$ (in)",
        zlabel="$z$ (in)",
    )
    ax.view_init(elev=24, azim=-126)
    ax.invert_yaxis()
    ax.set_zticks([94, 98, 102, 106, 110])
    ax.tick_params(axis="z", labelsize=8, pad=1)
    ax.set_box_aspect((1.8, 1.0, 0.42), zoom=1.16)
    ax.grid(True, alpha=0.35)
    ax.set_position([0.015, 0.025, 0.97, 0.95])
    save_figure(fig, "airfoil_sections_3d_spanwise")


def plot_stacked(x, sections, colours):
    # Keep even the thickest dimensional inboard profiles visually separate.
    spacing = 24.0
    offsets = np.arange(len(sections)) * spacing
    fig, ax = plt.subplots(figsize=(10.6, 15.5), constrained_layout=True)

    for section, colour, offset in zip(sections, colours, offsets):
        chord = section["chord_in"]
        x_in = x * chord / 100.0
        upper = section["upper"] * chord / 100.0 + offset
        lower = section["lower"] * chord / 100.0 + offset
        ax.fill_between(x_in, lower, upper, color=colour, alpha=0.18, linewidth=0)
        ax.plot(x_in, upper, color=colour, lw=1.7)
        ax.plot(x_in, lower, color=colour, lw=1.7)
        ax.axhline(offset, color=colour, lw=0.55, alpha=0.40)

    labels = [section["name"] for section in sections]
    ax.set_yticks(offsets, labels)
    ax.tick_params(axis="y", labelsize=8.6, pad=8)
    ax.set(
        title="Spanwise Evolution of the X-56A Airfoil Sections",
        xlabel="$x$ (in)",
        ylabel="Spanwise station  (profiles vertically separated)",
        xlim=(-1, 1.02 * max(section["chord_in"] for section in sections)),
        ylim=(-12, offsets[-1] + 12),
    )
    ax.set_aspect("equal", adjustable="box")
    ax.grid(axis="x")
    ax.grid(axis="y", visible=False)
    save_figure(fig, "airfoil_sections_spanwise_stacked")


def plot_geometry_metrics(x, sections, colours):
    fig, (ax_t, ax_c) = plt.subplots(
        2, 1, figsize=(12.5, 8.3), sharex=True, constrained_layout=True
    )

    for section, colour in zip(sections, colours):
        chord = section["chord_in"]
        x_in = x * chord / 100.0
        thickness = (section["upper"] - section["lower"]) * chord / 100.0
        camber = 0.5 * (section["upper"] + section["lower"]) * chord / 100.0
        label = section["name"]
        ax_t.plot(x_in, thickness, color=colour, lw=2.0, label=label)
        ax_c.plot(x_in, camber, color=colour, lw=2.0)

    ax_t.set_title("Local Thickness Distribution")
    ax_t.set_ylabel("Thickness, $t$ (in)")
    ax_t.set_ylim(bottom=0)
    ax_t.legend(
        ncol=3,
        loc="upper right",
        fontsize=8.5,
    )

    ax_c.axhline(0, color="#48515c", lw=0.8, alpha=0.75)
    ax_c.set_title("Mean Camber-Line Distribution")
    ax_c.set(
        xlabel="$x$ (in)",
        ylabel="Camber, $z_c$ (in)",
        xlim=(-1, 1.02 * max(section["chord_in"] for section in sections)),
    )
    fig.suptitle(
        "X-56A Section Geometry — Centreline to Outboard Wing",
        fontsize=16,
        fontweight="bold",
    )
    save_figure(fig, "airfoil_sections_thickness_and_camber")


def plot_combined_summary():
    """Combine the requested three finished figures on one A4 portrait page."""
    paths = [
        OUTPUT / "airfoil_sections_3d_spanwise.png",
        OUTPUT / "airfoil_sections_planform_positions.png",
        OUTPUT / "airfoil_sections_superimposed.png",
    ]
    images = []
    aspect_heights = []
    for path in paths:
        image = plt.imread(path)
        # Remove only uniform near-white outer margins.  Plot content, grids and
        # labels remain untouched.
        rgb = image[..., :3]
        content = np.any(rgb < 0.985, axis=2)
        rows, columns = np.where(content)
        pad = 18
        row_0 = max(0, int(rows.min()) - pad)
        row_1 = min(image.shape[0], int(rows.max()) + pad + 1)
        col_0 = max(0, int(columns.min()) - pad)
        col_1 = min(image.shape[1], int(columns.max()) + pad + 1)
        cropped = image[row_0:row_1, col_0:col_1]
        images.append(cropped)
        aspect_heights.append(cropped.shape[0] / cropped.shape[1])

    fig = plt.figure(figsize=(8.27, 11.69))
    grid = fig.add_gridspec(
        4,
        1,
        height_ratios=[
            0.82 * aspect_heights[0],
            aspect_heights[1],
            0.07,
            1.18 * aspect_heights[2],
        ],
    )
    axes = [fig.add_subplot(grid[0]), fig.add_subplot(grid[1]), fig.add_subplot(grid[3])]
    for ax, image in zip(axes, images):
        ax.imshow(image)
        ax.set_axis_off()
    fig.subplots_adjust(left=0.018, right=0.982, bottom=0.015, top=0.985, hspace=0.035)
    OUTPUT.mkdir(exist_ok=True)
    fig.savefig(
        OUTPUT / "airfoil_sections_combined_vertical.png",
        dpi=300,
        facecolor="white",
    )
    fig.savefig(
        OUTPUT / "airfoil_sections_combined_vertical.pdf",
        facecolor="white",
    )
    plt.close(fig)


def main():
    configure_style()
    x, sections = read_sections(SOURCE)
    colours = station_colours(sections)
    plot_superimposed(x, sections, colours)
    plot_planform_positions(sections, colours)
    plot_sections_3d(x, sections, colours)
    plot_stacked(x, sections, colours)
    plot_geometry_metrics(x, sections, colours)
    plot_combined_summary()
    print(f"Created 6 figures in PNG and PDF format under: {OUTPUT}")
    for path in sorted(OUTPUT.iterdir()):
        print(f"  {path.name}")


if __name__ == "__main__":
    main()
